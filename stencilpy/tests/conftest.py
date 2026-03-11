from __future__ import annotations

import datetime
from pathlib import Path

import openpyxl
import pytest
import yaml


@pytest.fixture
def tmp_dir(tmp_path: Path) -> Path:
    return tmp_path


@pytest.fixture
def sample_excel_v2(tmp_dir: Path) -> Path:
    """Create a sample Excel file matching the v2.0 schema."""
    path = tmp_dir / "lab_v2.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Discriminator
    ws["A1"] = "v2.0"

    # Patient info
    ws["B3"] = "Jane Doe"
    ws["B4"] = datetime.datetime(2024, 1, 15, 10, 30, 0)

    # First name / last name for computed fields
    ws["C3"] = "Jane"
    ws["C4"] = "Doe"

    # Readings (open-ended column D, starting at D5)
    ws["D5"] = 1.5
    ws["D6"] = 2.3
    ws["D7"] = 3.7
    ws["D8"] = 0.9
    # D9 is empty — signals end of open-ended range

    # Metadata (A10:B15)
    ws["A10"] = "lab_id"
    ws["B10"] = "LAB-001"
    ws["A11"] = "technician"
    ws["B11"] = "Dr. Smith"
    ws["A12"] = "method"
    ws["B12"] = "HPLC"

    # Results table (A20:D, with headers in row 20)
    ws["A20"] = "analyte"
    ws["A21"] = "Glucose"
    ws["A22"] = "Cholesterol"
    ws["B20"] = "value"
    ws["B21"] = 95.0
    ws["B22"] = 180.0
    ws["C20"] = "unit"
    ws["C21"] = "mg/dL"
    ws["C22"] = "mg/dL"
    ws["D20"] = "flag"
    ws["D21"] = "normal"
    ws["D22"] = "high"

    # Weight and height for BMI computed field
    ws["E3"] = 70.0  # weight in kg
    ws["E4"] = 1.75  # height in m

    ws.oddHeader.right.text = "v2.0-header"
    ws.oddFooter.center.text = "footer-note"

    wb.save(str(path))
    return path


@pytest.fixture
def sample_excel_v1(tmp_dir: Path) -> Path:
    """Create a sample Excel file matching the v1.0 schema."""
    path = tmp_dir / "lab_v1.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Discriminator
    ws["A1"] = "v1.0"

    # Patient name at A3
    ws["A3"] = "John Smith"

    # Readings in column C starting at C2
    ws["C2"] = 5.5
    ws["C3"] = 6.1
    ws["C4"] = 4.8

    # Sheet2 for results table
    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "Glucose"
    ws2["B1"] = 90.0
    ws2["C1"] = "mg/dL"
    ws2["D1"] = "normal"
    ws2["A2"] = "HbA1c"
    ws2["B2"] = 5.4
    ws2["C2"] = "%"
    ws2["D2"] = "normal"

    wb.save(str(path))
    return path


@pytest.fixture
def sample_excel_bad_disc(tmp_dir: Path) -> Path:
    """Create an Excel file with an unrecognized discriminator."""
    path = tmp_dir / "lab_bad.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "v99.0"
    wb.save(str(path))
    return path


@pytest.fixture
def sample_excel_header_disc(tmp_dir: Path) -> Path:
    """Create a workbook whose discriminator lives in the header."""
    path = tmp_dir / "lab_header_disc.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["B3"] = "Jane Doe"
    ws["B4"] = datetime.datetime(2024, 1, 15, 10, 30, 0)
    ws["D5"] = 1.5
    ws["D6"] = 2.3
    ws["E3"] = 70.0
    ws["E4"] = 1.75
    ws.oddHeader.right.text = "v2.0"
    wb.save(str(path))
    return path


@pytest.fixture
def sample_excel_no_disc_v2(tmp_dir: Path) -> Path:
    """Create a v2 workbook with no usable discriminator cell."""
    path = tmp_dir / "lab_v2_no_disc.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = ""
    ws["B3"] = "Jane Doe"
    ws["B4"] = datetime.datetime(2024, 1, 15, 10, 30, 0)
    ws["D5"] = 1.5
    ws["D6"] = 2.3
    ws["E3"] = 70.0
    ws["E4"] = 1.75

    wb.save(str(path))
    return path


@pytest.fixture
def ambiguous_schema_yaml(tmp_dir: Path) -> Path:
    """Write a schema whose versions are ambiguous without a discriminator."""
    path = tmp_dir / "ambiguous.stencil.yaml"
    schema = {
        "name": "ambiguous_report",
        "description": "Schema used to test ambiguous layout inference",
        "discriminator": {"cells": ["A1"]},
        "versions": {
            "v1": {"fields": {"shared": {"cell": "B2"}}},
            "v2": {"fields": {"shared": {"cell": "B2"}}},
        },
    }
    with open(path, "w") as f:
        yaml.dump(schema, f, default_flow_style=False)
    return path


@pytest.fixture
def ambiguous_excel_no_disc(tmp_dir: Path) -> Path:
    """Create a workbook that cannot be uniquely inferred."""
    path = tmp_dir / "ambiguous_no_disc.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = ""
    ws["B2"] = "shared value"
    wb.save(str(path))
    return path


@pytest.fixture
def sample_schema_dict() -> dict:
    """Return a sample schema as a Python dict."""
    return {
        "name": "lab_report",
        "description": "Monthly lab report from ACME Labs",
        "discriminator": {"cells": ["A1"]},
        "versions": {
            "v2.0": {
                "fields": {
                    "patient_name": {"cell": "B3"},
                    "sample_date": {"cell": "B4", "type": "datetime"},
                    "header_version": {"cell": "header:right"},
                    "footer_note": {"cell": "footer:center"},
                    "readings": {"range": "D5:D", "type": "list[float]"},
                    "results_table": {
                        "range": "A20:D",
                        "type": "table",
                    },
                    "metadata": {"range": "A10:B15", "type": "dict[str, str]"},
                    "weight": {"cell": "E3", "type": "float"},
                    "height": {"cell": "E4", "type": "float"},
                    "bmi": {"computed": "{weight} / ({height} ** 2)"},
                },
                "validation": {
                    "readings": {"min": 0, "max": 1000},
                    "patient_name": {"pattern": "^[A-Za-z ]+$"},
                },
            },
            "v1.0": {
                "fields": {
                    "patient_name": {"cell": "A3"},
                    "readings": {"range": "C2:C", "type": "list[float]"},
                    "results_table": {
                        "range": "Sheet2!A1:D",
                        "type": "table",
                        "columns": {"A": "analyte", "B": "value", "C": "unit", "D": "flag"},
                    },
                },
            },
        },
    }


@pytest.fixture
def sample_schema_yaml(tmp_dir: Path, sample_schema_dict: dict) -> Path:
    """Write a sample schema YAML file."""
    path = tmp_dir / "lab_report.stencil.yaml"
    with open(path, "w") as f:
        yaml.dump(sample_schema_dict, f, default_flow_style=False)
    return path


@pytest.fixture
def schema_dir(tmp_dir: Path, sample_schema_dict: dict) -> Path:
    """Create a directory with schema files."""
    schema_dir = tmp_dir / "schemas"
    schema_dir.mkdir()
    path = schema_dir / "lab_report.stencil.yaml"
    with open(path, "w") as f:
        yaml.dump(sample_schema_dict, f, default_flow_style=False)
    return schema_dir
