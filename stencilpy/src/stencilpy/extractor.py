from __future__ import annotations

import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from .addressing import CellAddress, RangeAddress, parse_cell, parse_range, _index_to_col
from .errors import StencilError
from .schema import FieldDef, ELEMENT_TYPE_MAP


def extract_fields(
    excel_path: str | Path,
    fields: dict[str, FieldDef],
    *,
    wb: openpyxl.Workbook | None = None,
) -> dict[str, Any]:
    """Extract all non-computed fields from an Excel file."""
    owned = wb is None
    if owned:
        wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
    try:
        result: dict[str, Any] = {}
        for name, field_def in fields.items():
            if field_def.is_computed:
                continue
            result[name] = _extract_field(wb, field_def)
        return result
    finally:
        if owned:
            wb.close()


def read_cell(excel_path: str | Path, cell_ref: str) -> Any:
    """Read a single cell value from an Excel file."""
    wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
    try:
        addr = parse_cell(cell_ref)
        ws = _get_sheet(wb, addr.sheet)
        return _read_cell_value(ws, addr.row, addr.col)
    finally:
        wb.close()


def _extract_field(wb: openpyxl.Workbook, field_def: FieldDef) -> Any:
    if field_def.cell:
        return _extract_cell(wb, field_def)
    elif field_def.range:
        return _extract_range(wb, field_def)
    else:
        raise ValueError(f"Field '{field_def.name}' has no cell or range")


def _extract_cell(wb: openpyxl.Workbook, field_def: FieldDef) -> Any:
    addr = parse_cell(field_def.cell)
    ws = _get_sheet(wb, addr.sheet)
    value = _read_cell_value(ws, addr.row, addr.col)
    return _coerce_scalar(value, field_def.resolved_type_str)


def _extract_range(wb: openpyxl.Workbook, field_def: FieldDef) -> Any:
    rng = parse_range(field_def.range)
    ws = _get_sheet(wb, rng.sheet)

    if field_def.is_table:
        return _extract_table(ws, rng, field_def)
    elif field_def.is_dict:
        return _extract_dict(ws, rng)
    elif field_def.is_list:
        return _extract_list(ws, rng, field_def)
    else:
        return _extract_list(ws, rng, field_def)


def _extract_list(ws: Worksheet, rng: RangeAddress, field_def: FieldDef) -> list[Any]:
    rows = _read_range_rows(ws, rng)
    element_type_str = field_def.resolved_type_str
    elem_type = ELEMENT_TYPE_MAP.get(element_type_str, str)

    result = []
    for row in rows:
        result.append(_coerce_value(row[0], elem_type))
    return result


def _extract_dict(ws: Worksheet, rng: RangeAddress) -> dict[str, str]:
    rows = _read_range_rows(ws, rng)
    result = {}
    for row in rows:
        if len(row) >= 2:
            key = str(row[0]) if row[0] is not None else ""
            val = str(row[1]) if row[1] is not None else ""
            result[key] = val
    return result


def _extract_table(
    ws: Worksheet, rng: RangeAddress, field_def: FieldDef
) -> list[dict[str, Any]]:
    rows = _read_range_rows(ws, rng)
    if not rows:
        return []

    if field_def.table_orientation == "vertical":
        return _extract_vertical_table(rows, rng, field_def.columns)

    if field_def.columns:
        headers = _build_column_headers(field_def.columns, rng)
        data_rows = rows
    else:
        headers = [str(v) if v is not None else f"col_{i}" for i, v in enumerate(rows[0])]
        data_rows = rows[1:]

    result = []
    for row in data_rows:
        record = {}
        for i, header in enumerate(headers):
            record[header] = row[i] if i < len(row) else None
        result.append(record)
    return result


def _extract_vertical_table(
    rows: list[list[Any]],
    rng: RangeAddress,
    columns: dict[str, str] | None = None,
) -> list[dict[str, Any]]:
    max_cols = max((len(row) for row in rows), default=0)
    if max_cols <= 1:
        return []

    headers = []
    for idx, row in enumerate(rows):
        row_key = str(rng.start_row + idx)
        if columns and row_key in columns:
            headers.append(columns[row_key])
        else:
            headers.append(str(row[0]) if row and row[0] is not None else f"row_{idx}")

    records: list[dict[str, Any]] = []
    for c in range(1, max_cols):
        record: dict[str, Any] = {}
        for r, row in enumerate(rows):
            key = headers[r]
            record[key] = row[c] if c < len(row) else None
        records.append(record)

    # Remove trailing records that are fully empty
    while records and all(v is None for v in records[-1].values()):
        records.pop()

    return records


def _build_column_headers(
    columns: dict[str, str], rng: RangeAddress
) -> list[str]:
    """Build ordered header list from explicit column mapping."""
    num_cols = rng.end_col - rng.start_col + 1
    headers = []
    for i in range(num_cols):
        col_letter = _index_to_col(rng.start_col + i)
        headers.append(columns.get(col_letter, col_letter))
    return headers


def _read_range_rows(ws: Worksheet, rng: RangeAddress) -> list[list[Any]]:
    """Read rows from a range. For open-ended ranges, read until first fully empty row."""
    if rng.end_row is not None:
        return _read_bounded_range(ws, rng)
    else:
        return _read_open_ended_range(ws, rng)


def _read_bounded_range(ws: Worksheet, rng: RangeAddress) -> list[list[Any]]:
    rows = []
    for row_idx in range(rng.start_row, rng.end_row + 1):
        row_data = []
        for col_idx in range(rng.start_col, rng.end_col + 1):
            row_data.append(_read_cell_value(ws, row_idx, col_idx))
        rows.append(row_data)
    # Strip trailing empty rows
    while rows and all(v is None for v in rows[-1]):
        rows.pop()
    return rows


def _read_open_ended_range(ws: Worksheet, rng: RangeAddress) -> list[list[Any]]:
    rows = []
    row_idx = rng.start_row
    max_row = ws.max_row or rng.start_row
    while row_idx <= max_row + 1:
        row_data = []
        for col_idx in range(rng.start_col, rng.end_col + 1):
            row_data.append(_read_cell_value(ws, row_idx, col_idx))
        if all(v is None for v in row_data):
            break
        rows.append(row_data)
        row_idx += 1
    return rows


def _read_cell_value(ws: Worksheet, row: int, col: int) -> Any:
    return ws.cell(row=row, column=col).value


def _get_sheet(wb: openpyxl.Workbook, sheet_name: str | None) -> Worksheet:
    if sheet_name is None:
        return wb.worksheets[0]
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
    return wb[sheet_name]


def _coerce_scalar(value: Any, type_str: str) -> Any:
    if value is None:
        return None
    type_map = {
        "str": str,
        "int": int,
        "float": float,
        "bool": bool,
        "datetime": lambda v: v if isinstance(v, datetime.datetime) else datetime.datetime.fromisoformat(str(v)),
        "date": lambda v: v.date() if isinstance(v, datetime.datetime) else v if isinstance(v, datetime.date) else datetime.date.fromisoformat(str(v)),
    }
    coercer = type_map.get(type_str)
    if coercer:
        return coercer(value)
    if type_str in {"any", "table", "dict[str, str]"} or type_str.startswith("list["):
        return value
    raise StencilError(f"Unknown scalar type '{type_str}'")


def _coerce_value(value: Any, target_type: type) -> Any:
    if value is None:
        return None
    if target_type == str:
        return str(value)
    if target_type == int:
        return int(value)
    if target_type == float:
        return float(value)
    if target_type == bool:
        return bool(value)
    return value
